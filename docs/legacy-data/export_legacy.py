"""Export all user tables from OvcinaLegacy SQL Server database to a single Excel file."""

import pyodbc
import openpyxl
import os
import sys

CONN_STR = (
    r"Driver={ODBC Driver 17 for SQL Server};"
    r"Server=(localdb)\OvcinaTemp;"
    r"Database=OvcinaLegacy;"
    r"Trusted_Connection=yes;"
)

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "ovcina-legacy-export.xlsx")
MAX_TEXT_LEN = 500


def main():
    print("Connecting to (localdb)\\OvcinaTemp / OvcinaLegacy ...")
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Get all user tables, excluding aspnet_* system tables
    cursor.execute("""
        SELECT TABLE_SCHEMA, TABLE_NAME
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_TYPE = 'BASE TABLE'
          AND TABLE_NAME NOT LIKE 'aspnet_%'
        ORDER BY TABLE_SCHEMA, TABLE_NAME
    """)
    tables = cursor.fetchall()

    if not tables:
        print("No user tables found!")
        sys.exit(1)

    print(f"Found {len(tables)} tables (excluding aspnet_* system tables).\n")

    wb = openpyxl.Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    summary = []

    for schema, table_name in tables:
        full_name = f"[{schema}].[{table_name}]"
        print(f"Exporting {full_name} ...", end=" ", flush=True)

        try:
            cursor.execute(f"SELECT * FROM {full_name}")
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()
        except Exception as e:
            print(f"ERROR: {e}")
            summary.append((schema, table_name, -1, str(e)))
            continue

        # Sheet name: max 31 chars for Excel, use table_name (prefix schema if needed)
        sheet_name = table_name[:31] if schema == "dbo" else f"{schema}.{table_name}"[:31]

        # Deduplicate sheet names if necessary
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = f"{schema}.{table_name}"[:31]
            counter = 2
            while sheet_name in existing:
                suffix = f"_{counter}"
                sheet_name = f"{schema}.{table_name}"[: 31 - len(suffix)] + suffix
                counter += 1

        ws = wb.create_sheet(title=sheet_name)

        # Write header row
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                # Truncate very long text fields
                if isinstance(value, str) and len(value) > MAX_TEXT_LEN:
                    value = value[:MAX_TEXT_LEN] + "... [TRUNCATED]"
                # Convert bytes to hex string for Excel compatibility
                if isinstance(value, (bytes, bytearray)):
                    value = value.hex()[:MAX_TEXT_LEN]
                ws.cell(row=row_idx, column=col_idx, value=value)

        row_count = len(rows)
        print(f"{row_count} rows, {len(columns)} columns")
        summary.append((schema, table_name, row_count, None))

    # Save workbook
    print(f"\nSaving to {OUTPUT_PATH} ...")
    wb.save(OUTPUT_PATH)
    print("Done!\n")

    # Print summary
    print("=" * 60)
    print("EXPORT SUMMARY")
    print("=" * 60)
    total_rows = 0
    for schema, table_name, row_count, error in summary:
        full = f"{schema}.{table_name}"
        if error:
            print(f"  {full:<40} ERROR: {error}")
        else:
            print(f"  {full:<40} {row_count:>6} rows")
            total_rows += row_count
    print("-" * 60)
    print(f"  {'TOTAL':<40} {total_rows:>6} rows")
    print(f"  Tables exported: {sum(1 for _, _, rc, _ in summary if rc >= 0)}")
    print(f"  Tables with errors: {sum(1 for _, _, _, e in summary if e)}")
    print(f"  Output: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
